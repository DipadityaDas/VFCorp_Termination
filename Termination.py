import os
import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Font


def create_excel_sheet() -> Worksheet:
	worksheet = wb.create_sheet('User Login IDs')
	del wb[wb.sheetnames[0]]
	worksheet['A1'] = 'User Name'
	worksheet['B1'] = 'User ID'
	worksheet['C1'] = 'Email'
	return worksheet


def cell_style(sheet: Worksheet) -> None:

	for row in sheet.iter_rows():
		for cell in row:
			cell.alignment = Alignment(horizontal='center', vertical='center')
			cell.font = Font(name='Cascadia Code')

	for column in sheet.columns:
		max_length = max(len(str(cell.value)) for cell in column)
		adjusted_width = (max_length + 2) * 1.2
		sheet.column_dimensions[column[0].column_letter].width = adjusted_width

	sheet.sheet_view.zoomScale = 180


def group_users(reports: list) -> None:
	print("All Termination Reports")
	print("-" * 70)

	for idx, file in enumerate(reports, start=1):
		print(f"{idx}.  {file}")

	print("=" * 70)

	data_frames = []
	for report in reports:
		df = pd.read_excel(dir_path + report, skiprows=1)
		data_frames.append(df)

	merged_column = pd.concat([df.iloc[:, 1] for df in data_frames], ignore_index=True)
	print("[INFO] Merged all the User Login IDs.")

	sorted_unique_column = sorted(merged_column.unique())
	print("[INFO] Deleted all duplicate User Login IDs and sorted the IDs.")

	for idx, data in enumerate(sorted_unique_column, start=2):
		ws.cell(row=idx, column=2).value = data

	cell_style(ws)

	print(f"[INFO] Found {len(sorted_unique_column)} User Login IDs.")
	print("[INFO] Successfully created the group.xlsx file.")


def convert_excel_to_csv():
	df = pd.read_excel(dir_path + excel_file)
	df.to_csv(dir_path + csv_file, index=False)
	print(f"[INFO] Conversion complete. Data saved to {csv_file}.")


if __name__ == '__main__':
	dir_path = "C:\\Users\\Dipaditya\\Downloads\\"
	excel_file = 'group.xlsx'
	csv_file = "Terminate_Users.csv"

	file_paths = [key for key in os.listdir(dir_path) if key.__contains__('Analytics')]

	print("=" * 70)
	print("Generating Log....")
	print("-" * 70)

	if file_paths:
		wb = Workbook()
		ws = create_excel_sheet()
		group_users(file_paths)
		wb.save(dir_path + excel_file)
		convert_excel_to_csv()
		os.remove(dir_path + excel_file)
	else:
		print('[INFO] No termination report found.')

